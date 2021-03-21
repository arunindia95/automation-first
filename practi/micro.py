import xlrd
import openpyxl
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
import time

driver = webdriver.Chrome(ChromeDriverManager().install())
driver.implicitly_wait(10)
driver.maximize_window()
driver.get("https://rarediseases.info.nih.gov/diseases/categories")
driver.execute_script("window.scrollBy(0,200)", "")
autoimmunie = driver.find_element(By.LINK_TEXT, "Autoimmune / Autoinflammatory diseases").click()
driver.execute_script("window.scrollBy(0,200)", "")
addision_disesase = driver.find_element(By.LINK_TEXT, "Addison's disease").click()
time.sleep(3)
scroll = driver.execute_script("window.scrollBy(0,600)", "")
summery = driver.find_element(By.XPATH, "//div[@id='panelContentOverview']").text
print(summery)
view = driver.find_element(By.XPATH, "//button[@class= 'btn btn-link phenotypeTableToggler']")
time.sleep(3)
scroll = driver.execute_script("arguments[0].scrollIntoView();", view)
view.click()
time.sleep(5)
medical_terms = driver.find_elements(By.XPATH, "//th[@class='phenotype-term-header']")
print(len(medical_terms))

medterms_list = driver.find_elements(By.XPATH, ".//*[@class='phenotype-term']")
length = len(medterms_list)

n = 0
# while n < len(medterms_list):

# for x in medterms_list:
#     # print(x[n].text)
#
#     if n % 3 == 0:
#         print(x.text)
#     n = n + 1

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "micro"
print(sheet.title)
sheet["A1"].value = "category"
sheet["B1"].value = "desease"
sheet["C1"].value = "summery"
sheet["D1"].value = "medical terms"

# nfc_east = ('DAL', 'WAS', 'PHI', 'NYG')
# wb = Workbook()
# ws = wb.active
row_cell = 2
column_cell=2
# for i in nfc_east:
for x in medterms_list:
    # print(x[n].text)

    if n % 3 == 0:
        print(x.text)
    n = n + 1

    # column_cell = 'D'
    sheet.cell(row=row_cell, column=column_cell).value = x
    row_cell = row_cell + 1
workbook.save("C:\\Users\\Admin\\Documents\\micro.xlsx")

time.sleep(4)
driver.quit()
